#!/usr/bin/env python3
"""
COMPLETE INTEGRATED EXCEL ANALYZER + CHATBOT - VS CODE VERSION V2
- Upload Excel file (.xlsx) → Analyzes and creates chatbot
- Upload .txt file (pre-analyzed) → Directly loads into chatbot
- Handles 4M+ rows
- Interactive AI chatbot

INSTRUCTIONS:
1. Install dependencies: pip install openpyxl openai python-dotenv
2. Create .env file with: OPENAI_API_KEY=your_key_here
3. Run: python excel_analyzer_chatbot.py
4. Provide Excel (.xlsx) OR pre-analyzed text file (.txt) path when prompted
5. Ask questions!
"""
!pip install openpyxl openai python-dotenv

import openpyxl
from datetime import datetime
import os
from decimal import Decimal, InvalidOperation
import tempfile
import json
import sqlite3
import re
import time
from typing import Dict, List, Optional, Tuple
from openai import OpenAI
from pathlib import Path

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
    print("✅ .env file loaded")
except ImportError:
    print("⚠️  python-dotenv not installed. Using environment variables directly.")

print("✅ Libraries imported successfully!")

# ========================================
# CONFIGURATION
# ========================================

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")

if OPENAI_API_KEY:
    print("✅ OpenAI API key loaded from environment")
    client = OpenAI(api_key=OPENAI_API_KEY)
else:
    print("❌ OPENAI_API_KEY not found in environment")
    print("   Create .env file with: OPENAI_API_KEY=your_key_here")
    print("   Or set environment variable: export OPENAI_API_KEY=your_key_here")
    client = None

# Global variables
DB_CONN = None
SUMMARY_TEXT = ""
SUMMARY_DATA = {}
TABLE_NAME = "data"
FILE_METADATA = {}

# ========================================
# EXCEL ANALYZER FUNCTIONS
# ========================================

def extract_metadata(ws, max_search_rows=50):
    """Extract company/institution details"""
    metadata_dict = {}
    metadata_lines = []
    
    for row_idx in range(1, min(max_search_rows + 1, 100)):
        try:
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            non_empty_cells = [cell for cell in row if cell is not None and str(cell).strip() != '']
            
            if len(non_empty_cells) >= 5:
                has_text = any(isinstance(cell, str) for cell in non_empty_cells)
                if has_text:
                    return metadata_dict, metadata_lines, row_idx
            
            if 1 <= len(non_empty_cells) <= 4:
                first_cell = str(non_empty_cells[0]).strip()
                
                if ':' in first_cell:
                    parts = first_cell.split(':', 1)
                    key = parts[0].strip()
                    value = parts[1].strip() if len(parts) > 1 else ''
                    metadata_dict[key] = value
                    metadata_lines.append(first_cell)
                elif len(non_empty_cells) == 2:
                    key = str(non_empty_cells[0]).strip()
                    value = str(non_empty_cells[1]).strip()
                    metadata_dict[key] = value
                    metadata_lines.append(f"{key}: {value}")
                else:
                    metadata_lines.append(first_cell)
        except Exception:
            continue
    
    return metadata_dict, metadata_lines, None


def is_numeric_value(value):
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.replace(',', '').replace('$', '').replace('€', '').replace('₹', '').strip())
            return True
        except:
            return False
    return False


def safe_numeric_conversion(value):
    if value is None:
        return None
    try:
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        elif isinstance(value, str):
            cleaned = value.replace(',', '').replace('$', '').replace('€', '').replace('₹', '').strip()
            return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    return None


def is_numeric_column(column_data, threshold=0.6):
    numeric_count = 0
    total_count = 0
    
    for value in column_data:
        if value is not None and str(value).strip() != '':
            total_count += 1
            if is_numeric_value(value):
                numeric_count += 1
    
    if total_count == 0:
        return False
    return (numeric_count / total_count) >= threshold


def is_summary_row(row, headers=None):
    summary_keywords = [
        'NET TOTAL', 'GRAND TOTAL', 'TOTAL', 'SUBTOTAL', 'SUM', 
        'NET', 'GROSS TOTAL', 'OVERALL', 'AGGREGATE', 'CONSOLIDATED'
    ]
    
    for cell in row:
        if cell and isinstance(cell, str):
            upper_cell = str(cell).upper().strip()
            for keyword in summary_keywords:
                if upper_cell == keyword or upper_cell.startswith(keyword + ' '):
                    return True, upper_cell
    return False, None


def extract_headers(ws, header_row_idx):
    header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    headers = []
    
    for idx, cell in enumerate(header_row):
        if cell is not None and str(cell).strip() != '':
            headers.append(str(cell).strip())
        else:
            break
    return headers


def analyze_excel_optimized(excel_file_path):
    """Analyze Excel file"""
    
    print(f"📂 Reading: {excel_file_path}")
    
    wb = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)
    ws = wb.active
    
    metadata_dict, metadata_lines, header_row_idx = extract_metadata(ws)
    
    if header_row_idx is None:
        print("❌ ERROR: Could not detect header")
        wb.close()
        return None, None, None
    
    print(f"✅ Header at row {header_row_idx}")
    
    headers = extract_headers(ws, header_row_idx)
    num_columns = len(headers)
    
    print(f"📊 {num_columns} columns")
    
    # Sample data
    data_start_row = header_row_idx + 1
    sample_data = {i: [] for i in range(num_columns)}
    
    for row in ws.iter_rows(min_row=data_start_row, max_row=data_start_row + 200, values_only=True):
        is_summary, _ = is_summary_row(row, headers)
        if is_summary:
            break
        for idx in range(num_columns):
            if idx < len(row):
                sample_data[idx].append(row[idx])
    
    numeric_columns = {}
    for idx, header in enumerate(headers):
        if is_numeric_column(sample_data[idx]):
            numeric_columns[idx] = header
    
    print(f"✅ {len(numeric_columns)} numeric columns")
    
    wb.close()
    wb = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)
    ws = wb.active
    
    stats = {}
    col_idx_to_key = {}
    
    for col_idx, col_name in numeric_columns.items():
        unique_key = f"{col_name}_COL{col_idx}"
        col_idx_to_key[col_idx] = unique_key
        stats[unique_key] = {
            'column_name': col_name,
            'column_index': col_idx,
            'total': Decimal('0'),
            'count': 0,
            'highest': {'value': Decimal('-Infinity'), 'row_data': {}, 'row_number': None},
            'lowest': {'value': Decimal('Infinity'), 'row_data': {}, 'row_number': None}
        }
    
    temp_data_file = tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', delete=False, suffix='.tmp')
    temp_data_path = temp_data_file.name
    
    for line in metadata_lines:
        temp_data_file.write(line + '\n')
    if metadata_lines:
        temp_data_file.write('\n')
    temp_data_file.write('\t'.join(headers) + '\n')
    
    print(f"⏳ Processing rows...")
    
    row_count = 0
    current_excel_row = data_start_row
    summary_rows_found = []
    
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        is_summary, summary_text = is_summary_row(row, headers)
        if is_summary:
            summary_rows_found.append((current_excel_row, summary_text))
            break
        
        row_data = []
        current_row_dict = {}
        original_values = {}
        has_any_data = False
        
        for col_idx in range(num_columns):
            cell_value = None
            if col_idx < len(row):
                cell_value = row[col_idx]
            
            original_values[col_idx] = cell_value
            
            display_value = ''
            if cell_value is not None:
                if isinstance(cell_value, datetime):
                    display_value = cell_value.strftime('%d-%m-%Y')
                else:
                    display_value = str(cell_value).strip()
                if display_value != '':
                    has_any_data = True
            
            row_data.append(display_value)
            current_row_dict[headers[col_idx]] = display_value
        
        if not has_any_data:
            current_excel_row += 1
            continue
        
        row_count += 1
        
        for col_idx in numeric_columns.keys():
            cell_value = original_values.get(col_idx)
            if cell_value is not None:
                numeric_val = safe_numeric_conversion(cell_value)
                if numeric_val is not None:
                    unique_key = col_idx_to_key[col_idx]
                    stats[unique_key]['total'] += numeric_val
                    stats[unique_key]['count'] += 1
                    
                    if stats[unique_key]['highest']['value'] == Decimal('-Infinity') or numeric_val > stats[unique_key]['highest']['value']:
                        stats[unique_key]['highest']['value'] = numeric_val
                        stats[unique_key]['highest']['row_number'] = current_excel_row
                        stats[unique_key]['highest']['row_data'] = current_row_dict.copy()
                    
                    if stats[unique_key]['lowest']['value'] == Decimal('Infinity') or numeric_val < stats[unique_key]['lowest']['value']:
                        stats[unique_key]['lowest']['value'] = numeric_val
                        stats[unique_key]['lowest']['row_number'] = current_excel_row
                        stats[unique_key]['lowest']['row_data'] = current_row_dict.copy()
        
        temp_data_file.write('\t'.join(row_data) + '\n')
        
        if row_count % 10000 == 0:
            print(f"   {row_count:,} rows...")
        
        current_excel_row += 1
    
    temp_data_file.close()
    wb.close()
    
    print(f"✅ Processed {row_count:,} rows")
    
    # Generate summary
    summary_lines = []
    summary_lines.append("=" * 100)
    summary_lines.append("EXCEL ANALYSIS SUMMARY")
    summary_lines.append("=" * 100)
    summary_lines.append(f"File: {excel_file_path}")
    summary_lines.append(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    summary_lines.append("")
    
    if metadata_dict:
        summary_lines.append("COMPANY/INSTITUTION DETAILS:")
        summary_lines.append("-" * 100)
        for key, value in metadata_dict.items():
            summary_lines.append(f"{key}: {value}")
        summary_lines.append("")
    
    summary_lines.append("DATA SUMMARY:")
    summary_lines.append("-" * 100)
    summary_lines.append(f"Total Data Rows Processed: {row_count:,}")
    summary_lines.append(f"Total Columns: {num_columns}")
    summary_lines.append(f"Numeric Columns Detected: {len(numeric_columns)}")
    summary_lines.append(f"Column Names: {', '.join(headers)}")
    
    if summary_rows_found:
        summary_lines.append(f"\n⚠️  SUMMARY ROWS DETECTED:")
        for row_num, text in summary_rows_found:
            summary_lines.append(f"   - Row {row_num}: {text}")
    
    summary_lines.append("")
    summary_lines.append("=" * 100)
    summary_lines.append("NUMERIC COLUMNS ANALYSIS:")
    summary_lines.append("=" * 100)
    
    sorted_keys = sorted(stats.keys(), key=lambda k: stats[k]['column_index'])
    
    for unique_key in sorted_keys:
        col_stats = stats[unique_key]
        col_name = col_stats['column_name']
        col_idx = col_stats['column_index']
        
        summary_lines.append("")
        summary_lines.append(f"📊 {col_name.upper()} (Column {col_idx + 1})")
        summary_lines.append("-" * 100)
        
        if col_stats['count'] > 0:
            total_val = float(col_stats['total'])
            summary_lines.append(f"Total: {total_val:,.2f}")
            summary_lines.append(f"Count: {col_stats['count']:,}")
            average = total_val / col_stats['count']
            summary_lines.append(f"Average: {average:,.2f}")
            
            if col_stats['highest']['value'] != Decimal('-Infinity'):
                highest_val = float(col_stats['highest']['value'])
                summary_lines.append(f"\n🔼 HIGHEST {col_name}: {highest_val:,.2f}")
                summary_lines.append(f"   Excel Row Number: {col_stats['highest']['row_number']}")
                summary_lines.append(f"   Complete Row Data:")
                for header in headers:
                    value = col_stats['highest']['row_data'].get(header, '')
                    summary_lines.append(f"      {header}: {value}")
            
            if col_stats['lowest']['value'] != Decimal('Infinity'):
                lowest_val = float(col_stats['lowest']['value'])
                summary_lines.append(f"\n🔽 LOWEST {col_name}: {lowest_val:,.2f}")
                summary_lines.append(f"   Excel Row Number: {col_stats['lowest']['row_number']}")
                summary_lines.append(f"   Complete Row Data:")
                for header in headers:
                    value = col_stats['lowest']['row_data'].get(header, '')
                    summary_lines.append(f"      {header}: {value}")
    
    summary_lines.append("")
    summary_lines.append("=" * 100)
    summary_lines.append("END OF ANALYSIS")
    summary_lines.append("=" * 100)
    summary_lines.append("")
    summary_lines.append("COMPLETE DATA TABLE:")
    summary_lines.append("-" * 100)
    
    summary_text = '\n'.join(summary_lines)
    
    return summary_text, temp_data_path, {
        'metadata': metadata_dict,
        'row_count': row_count,
        'headers': headers,
        'numeric_columns': list(numeric_columns.values())
    }


# ========================================
# SUMMARY PARSER
# ========================================

def parse_all_summary_values(summary_text: str) -> Dict:
    """Extract all calculations AND metadata"""
    
    calculations = {}
    lines = summary_text.split('\n')
    
    # Extract metadata from header section
    in_metadata = False
    for line in lines:
        line_stripped = line.strip()
        
        # Check for company/institution details section
        if 'COMPANY/INSTITUTION DETAILS:' in line_stripped.upper():
            in_metadata = True
            continue
        
        # Check for data summary section (end of metadata)
        if 'DATA SUMMARY:' in line_stripped.upper():
            in_metadata = False
            continue
        
        # Extract metadata key-value pairs
        if in_metadata and ':' in line_stripped and not line_stripped.startswith('-'):
            parts = line_stripped.split(':', 1)
            if len(parts) == 2:
                key = parts[0].strip().upper().replace(' ', '_').replace('/', '_')
                value = parts[1].strip()
                calculations[key.lower()] = value
    
    current_column = None
    current_extreme = None
    current_extreme_column = None
    current_row_data = {}
    capturing_row = False
    
    for line in lines:
        line_stripped = line.strip()
        
        if not line_stripped:
            if capturing_row and current_row_data and current_extreme and current_extreme_column:
                key = f"{current_extreme}_{current_extreme_column.lower()}_data"
                calculations[key] = current_row_data.copy()
                current_row_data = {}
                capturing_row = False
            continue
        
        match = re.search(r'📊\s+([A-Za-z0-9_\s\.]+?)\s+\(Column\s+\d+\)', line_stripped)
        if match:
            current_column = match.group(1).strip().replace(' ', '_').replace('.', '_').upper()
            current_extreme = None
            capturing_row = False
            continue
        
        if current_column:
            match = re.search(r'^\s*Total:\s*([\d,]+\.?\d*)', line_stripped, re.IGNORECASE)
            if match:
                value = match.group(1).replace(',', '')
                calculations[f"total_{current_column.lower()}"] = float(value)
                continue
            
            match = re.search(r'^\s*Count:\s*([\d,]+)', line_stripped, re.IGNORECASE)
            if match:
                value = match.group(1).replace(',', '')
                calculations[f"count_{current_column.lower()}"] = int(value)
                continue
            
            match = re.search(r'^\s*Average:\s*([\d,]+\.?\d*)', line_stripped, re.IGNORECASE)
            if match:
                value = match.group(1).replace(',', '')
                calculations[f"average_{current_column.lower()}"] = float(value)
                continue
        
        match = re.search(r'🔼\s*HIGHEST\s+([A-Za-z0-9_\s\.]+?):\s*([\d,]+\.?\d*)', line_stripped, re.IGNORECASE)
        if match:
            column = match.group(1).strip().replace(' ', '_').replace('.', '_').upper()
            value = match.group(2).replace(',', '')
            calculations[f"highest_{column.lower()}"] = float(value)
            current_extreme = "highest"
            current_extreme_column = column
            current_row_data = {}
            capturing_row = False
            continue
        
        match = re.search(r'🔽\s*LOWEST\s+([A-Za-z0-9_\s\.]+?):\s*([\d,]+\.?\d*)', line_stripped, re.IGNORECASE)
        if match:
            column = match.group(1).strip().replace(' ', '_').replace('.', '_').upper()
            value = match.group(2).replace(',', '')
            calculations[f"lowest_{column.lower()}"] = float(value)
            current_extreme = "lowest"
            current_extreme_column = column
            current_row_data = {}
            capturing_row = False
            continue
        
        if current_extreme and current_extreme_column:
            if 'complete row data' in line_stripped.lower():
                capturing_row = True
                continue
            
            if capturing_row:
                match = re.search(r'^\s*([A-Za-z0-9_\s\.]+?):\s*(.+)$', line_stripped)
                if match:
                    field = match.group(1).strip().replace(' ', '_').upper()
                    value = match.group(2).strip()
                    current_row_data[field] = value
                    continue
        
        match = re.search(r'Total\s+Data\s+Rows?(?:\s+Processed)?:\s*([\d,]+)', line_stripped, re.IGNORECASE)
        if match:
            value = match.group(1).replace(',', '')
            calculations["total_data_rows"] = int(value)
            calculations["total_rows"] = int(value)
            calculations["row_count"] = int(value)
        
        # Extract File path/name
        match = re.search(r'File:\s*(.+)', line_stripped, re.IGNORECASE)
        if match:
            calculations["file_name"] = match.group(1).strip()
        
        # Extract Analysis Date
        match = re.search(r'Analysis\s+Date:\s*(.+)', line_stripped, re.IGNORECASE)
        if match:
            calculations["analysis_date"] = match.group(1).strip()
        
        # Extract Column Names
        match = re.search(r'Column\s+Names:\s*(.+)', line_stripped, re.IGNORECASE)
        if match:
            calculations["column_names"] = match.group(1).strip()
        
        # Extract Total Columns
        match = re.search(r'Total\s+Columns:\s*(\d+)', line_stripped, re.IGNORECASE)
        if match:
            calculations["total_columns"] = int(match.group(1))
        
        # Extract Numeric Columns count
        match = re.search(r'Numeric\s+Columns\s+Detected:\s*(\d+)', line_stripped, re.IGNORECASE)
        if match:
            calculations["numeric_columns_count"] = int(match.group(1))
    
    print(f"✅ Extracted {len(calculations)} values")
    return calculations


# ========================================
# LOAD FROM TXT FILE (PRE-ANALYZED)
# ========================================

def load_from_txt_file(txt_file_path: str) -> Tuple[str, str]:
    """Load pre-analyzed .txt file and separate summary from data"""
    
    print(f"📂 Loading pre-analyzed file: {txt_file_path}")
    
    with open(txt_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    lines = content.split('\n')
    
    # Find where data table starts
    data_start_idx = None
    for i, line in enumerate(lines):
        if 'COMPLETE DATA TABLE:' in line.upper():
            data_start_idx = i + 2  # Skip the separator line
            break
    
    if data_start_idx is None:
        # Try to find by tab count
        for i, line in enumerate(lines):
            if line.count('\t') >= 5:
                data_start_idx = i
                break
    
    if data_start_idx is None:
        raise ValueError("Could not find data table in file")
    
    # Split summary and data
    summary_text = '\n'.join(lines[:data_start_idx])
    data_text = '\n'.join(lines[data_start_idx:])
    
    # Create temp file with data
    temp_data_file = tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', delete=False, suffix='.tmp')
    temp_data_path = temp_data_file.name
    temp_data_file.write(data_text)
    temp_data_file.close()
    
    print(f"✅ Summary and data separated")
    
    return summary_text, temp_data_path


# ========================================
# DATABASE LOADING
# ========================================

def clean_column_name(name: str) -> str:
    if not name or not name.strip():
        return 'column'
    name = re.sub(r'[^\w\s-]', '', name)
    name = re.sub(r'[-\s]+', '_', name)
    name = name.strip('_')
    return name if name else 'column'


def load_data_from_temp_file(temp_file_path: str, summary_text: str) -> Dict:
    """Load data into SQLite"""
    global DB_CONN, SUMMARY_TEXT, SUMMARY_DATA, TABLE_NAME, FILE_METADATA
    
    print(f"\n📂 Loading data...")
    
    with open(temp_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    lines = content.split('\n')
    
    # Find table start
    data_start_idx = 0
    for i, line in enumerate(lines):
        if line.count('\t') >= 5:
            data_start_idx = i
            break
    
    data_lines = [l for l in lines[data_start_idx:] if l.strip()]
    
    # Extract headers
    header_line = data_lines[0]
    raw_headers = [h.strip() for h in header_line.split('\t') if h.strip()]
    headers = [clean_column_name(h) for h in raw_headers]
    
    # Make unique
    seen = {}
    for i in range(len(headers)):
        if headers[i] in seen:
            seen[headers[i]] += 1
            headers[i] = f"{headers[i]}_{seen[headers[i]]}"
        else:
            seen[headers[i]] = 0
    
    num_columns = len(headers)
    
    # Extract rows
    rows = []
    for line in data_lines[1:]:
        if not line.strip():
            continue
        values = [v.strip() for v in line.split('\t')]
        if len(values) < num_columns * 0.5:
            continue
        if len(values) < num_columns:
            values.extend([''] * (num_columns - len(values)))
        elif len(values) > num_columns:
            values = values[:num_columns]
        rows.append(values)
    
    print(f"✅ {len(rows):,} rows loaded")
    
    # Parse summary
    SUMMARY_TEXT = summary_text
    SUMMARY_DATA = parse_all_summary_values(summary_text)
    
    FILE_METADATA = {
        'data_rows': len(rows),
        'columns': len(headers),
        'summary_calculations': len(SUMMARY_DATA)
    }
    
    # Create database
    print(f"⏳ Creating database...")
    
    DB_CONN = sqlite3.connect(':memory:')
    DB_CONN.execute('PRAGMA journal_mode = OFF')
    DB_CONN.execute('PRAGMA synchronous = OFF')
    
    cursor = DB_CONN.cursor()
    columns_def = ', '.join([f'"{h}" TEXT' for h in headers])
    cursor.execute(f'CREATE TABLE {TABLE_NAME} ({columns_def})')
    
    insert_query = f'INSERT INTO {TABLE_NAME} VALUES ({",".join(["?"] * len(headers))})'
    cursor.executemany(insert_query, rows)
    DB_CONN.commit()
    
    print(f"✅ Database ready!")
    
    return {"headers": headers, "total_rows": len(rows), "file_metadata": FILE_METADATA}


# ========================================
# QUERY FUNCTIONS
# ========================================

def get_summary() -> Dict:
    return {"found": True, "summary": SUMMARY_TEXT, "all_values": SUMMARY_DATA, "file_metadata": FILE_METADATA}

def get_metadata(info_type: str = "all") -> Dict:
    """Get file/company metadata like company name, address, row count, etc."""
    metadata = {}
    
    # Collect all metadata-related keys
    metadata_keys = [k for k in SUMMARY_DATA.keys() if not any(x in k for x in ['total_', 'count_', 'average_', 'highest_', 'lowest_']) or k in ['total_data_rows', 'total_rows', 'row_count', 'total_columns']]
    
    for key in metadata_keys:
        metadata[key] = SUMMARY_DATA[key]
    
    if info_type == "all":
        return {"found": True, "metadata": metadata}
    else:
        # Try to find specific info
        info_type_normalized = info_type.lower().replace(' ', '_')
        if info_type_normalized in SUMMARY_DATA:
            return {"found": True, "info_type": info_type, "value": SUMMARY_DATA[info_type_normalized]}
        
        # Partial match
        matches = {k: v for k, v in SUMMARY_DATA.items() if info_type_normalized in k}
        if matches:
            return {"found": True, "info_type": info_type, "matches": matches}
        
        return {"found": False, "error": f"'{info_type}' not found in metadata"}

def get_value_from_summary(query_type: str, column: str = None) -> Dict:
    if not column:
        return {"found": False, "error": "Column required"}
    
    column_normalized = column.upper().replace(' ', '_').replace('.', '_')
    lookup_key = f"{query_type.lower()}_{column_normalized.lower()}"
    
    if lookup_key in SUMMARY_DATA:
        return {"found": True, "query_type": query_type, "column": column, "value": SUMMARY_DATA[lookup_key]}
    
    return {"found": False, "error": f"{query_type} {column} not found"}

def get_extreme_row_data(query_type: str, column: str) -> Dict:
    if query_type not in ['highest', 'lowest']:
        return {"found": False, "error": "Use 'highest' or 'lowest'"}
    
    column_normalized = column.upper().replace(' ', '_').replace('.', '_')
    lookup_key = f"{query_type.lower()}_{column_normalized.lower()}_data"
    
    if lookup_key in SUMMARY_DATA:
        return {"found": True, "query_type": query_type, "column": column, "row_data": SUMMARY_DATA[lookup_key]}
    
    return {"found": False, "error": f"{query_type} {column} data not found"}

def get_column_names() -> Dict:
    cursor = DB_CONN.cursor()
    cursor.execute(f"PRAGMA table_info({TABLE_NAME})")
    columns = [row[1] for row in cursor.fetchall()]
    return {"found": True, "columns": columns, "total_columns": len(columns)}

def get_rows(n: int = 10, offset: int = 0) -> Dict:
    cursor = DB_CONN.cursor()
    cursor.execute(f"SELECT * FROM {TABLE_NAME} LIMIT ? OFFSET ?", (n, offset))
    columns = [desc[0] for desc in cursor.description]
    rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
    return {"found": True, "count": len(rows), "offset": offset, "rows": rows}

def search_by_column(column: str, value: str, limit: int = 100) -> Dict:
    cursor = DB_CONN.cursor()
    cursor.execute(f"PRAGMA table_info({TABLE_NAME})")
    valid_columns = [row[1] for row in cursor.fetchall()]
    
    if column not in valid_columns:
        matches = [c for c in valid_columns if column.lower() in c.lower()]
        if matches:
            column = matches[0]
        else:
            return {"found": False, "error": f"Column '{column}' not found"}
    
    query = f'SELECT * FROM {TABLE_NAME} WHERE "{column}" LIKE ? LIMIT ?'
    cursor.execute(query, (f'%{value}%', limit))
    columns = [desc[0] for desc in cursor.description]
    rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
    return {"found": len(rows) > 0, "count": len(rows), "results": rows}

def list_all_summary_values() -> Dict:
    return {"found": True, "total_values": len(SUMMARY_DATA), "all_values": SUMMARY_DATA}


# ========================================
# CHATBOT AGENT
# ========================================

TOOLS = [
    {"type": "function", "function": {"name": "get_summary", "description": "Get full summary text", "parameters": {"type": "object", "properties": {}, "required": []}}},
    {"type": "function", "function": {"name": "get_metadata", "description": "Get company/file metadata (company name, address, row count, etc.)", "parameters": {"type": "object", "properties": {"info_type": {"type": "string", "description": "Type of info: 'all', 'company_name', 'address', 'total_rows', etc."}}, "required": []}}},
    {"type": "function", "function": {"name": "get_value_from_summary", "description": "Get pre-calculated value", "parameters": {"type": "object", "properties": {"query_type": {"type": "string", "enum": ["total", "highest", "lowest", "average", "count"]}, "column": {"type": "string"}}, "required": ["query_type", "column"]}}},
    {"type": "function", "function": {"name": "get_extreme_row_data", "description": "Get complete row data", "parameters": {"type": "object", "properties": {"query_type": {"type": "string", "enum": ["highest", "lowest"]}, "column": {"type": "string"}}, "required": ["query_type", "column"]}}},
    {"type": "function", "function": {"name": "list_all_summary_values", "description": "List all values", "parameters": {"type": "object", "properties": {}, "required": []}}},
    {"type": "function", "function": {"name": "get_column_names", "description": "Get columns", "parameters": {"type": "object", "properties": {}, "required": []}}},
    {"type": "function", "function": {"name": "get_rows", "description": "Get rows", "parameters": {"type": "object", "properties": {"n": {"type": "integer", "default": 10}, "offset": {"type": "integer", "default": 0}}, "required": []}}},
    {"type": "function", "function": {"name": "search_by_column", "description": "Search rows", "parameters": {"type": "object", "properties": {"column": {"type": "string"}, "value": {"type": "string"}, "limit": {"type": "integer", "default": 100}}, "required": ["column", "value"]}}}
]

TOOL_MAP = {
    "get_summary": get_summary,
    "get_metadata": get_metadata,
    "get_value_from_summary": get_value_from_summary,
    "get_extreme_row_data": get_extreme_row_data,
    "list_all_summary_values": list_all_summary_values,
    "get_column_names": get_column_names,
    "get_rows": get_rows,
    "search_by_column": search_by_column
}

def query_agent(question: str, max_iterations: int = 15) -> str:
    if not client:
        return "❌ OpenAI client not initialized"
    
    system_prompt = f"""You are a data assistant. Answer ONLY from pre-calculated values.

AVAILABLE DATA:
- Company/Institution details (company_name, address, etc.)
- File metadata (file_name, analysis_date, total_rows, total_columns)
- Pre-calculated statistics (totals, averages, highest, lowest)
- Complete row data for extreme values
- Searchable database with all rows

RULES:
1. ALL calculations are PRE-CALCULATED - NEVER calculate manually
2. For company details: use list_all_summary_values() to see what's available
3. For row count: use get_value_from_summary(query_type="total_rows")
4. For numeric totals/averages: use get_value_from_summary()
5. For complete rows: use get_extreme_row_data()
6. To see all available values: use list_all_summary_values()

Total pre-calculated values available: {len(SUMMARY_DATA)}
"""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": question}
    ]
    
    iteration = 0
    
    while iteration < max_iterations:
        iteration += 1
        
        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=messages,
                tools=TOOLS,
                tool_choice="auto",
                temperature=0
            )
            
            assistant_message = response.choices[0].message
            messages.append(assistant_message)
            
            if not assistant_message.tool_calls:
                return assistant_message.content or "No response."
            
            for tool_call in assistant_message.tool_calls:
                tool_name = tool_call.function.name
                tool_args = json.loads(tool_call.function.arguments)
                
                try:
                    result = TOOL_MAP[tool_name](**tool_args)
                    result_str = json.dumps(result, indent=2, ensure_ascii=False)
                    if len(result_str) > 12000:
                        result_str = result_str[:12000] + "\n...[truncated]"
                    
                    messages.append({
                        "role": "tool",
                        "tool_call_id": tool_call.id,
                        "content": result_str
                    })
                except Exception as e:
                    messages.append({
                        "role": "tool",
                        "tool_call_id": tool_call.id,
                        "content": f"ERROR: {str(e)}"
                    })
        except Exception as e:
            return f"Error: {str(e)}"
    
    return "⚠️ Max iterations"


# ========================================
# FILE INPUT HELPER
# ========================================

def get_file_path():
    """Get file path from user with validation"""
    while True:
        filepath = input("\n📁 Enter file path (.xlsx or .txt): ").strip()
        
        # Remove quotes if present
        filepath = filepath.strip('"').strip("'")
        
        if not filepath:
            print("❌ No path entered. Try again.")
            continue
        
        # Expand user home directory
        filepath = os.path.expanduser(filepath)
        
        # Check if file exists
        if not os.path.exists(filepath):
            print(f"❌ File not found: {filepath}")
            print("   Please check the path and try again.")
            continue
        
        # Check file extension
        file_ext = os.path.splitext(filepath)[1].lower()
        if file_ext not in ['.xlsx', '.xlsm', '.xltx', '.xltm', '.txt']:
            print(f"❌ Unsupported file type: {file_ext}")
            print("   Please provide .xlsx or .txt file.")
            continue
        
        return filepath


# ========================================
# MAIN FUNCTION
# ========================================

def main():
    
    if not OPENAI_API_KEY:
        print("\n❌ ERROR: OPENAI_API_KEY not set")
        print("\nSetup options:")
        print("1. Create .env file:")
        print("   echo 'OPENAI_API_KEY=your_key_here' > .env")
        print("2. Set environment variable:")
        print("   export OPENAI_API_KEY=your_key_here")
        return
    
    filepath = get_file_path()
    file_ext = os.path.splitext(filepath)[1].lower()
    
    try:
        if file_ext == '.txt':
            # Pre-analyzed file
            print("\n" + "="*80)
            print("MODE: LOADING PRE-ANALYZED FILE")
            print("="*80 + "\n")
            
            start = time.time()
            summary_text, temp_data_path = load_from_txt_file(filepath)
            print(f"✅ Loaded ({time.time() - start:.2f}s)")
            
        elif file_ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            # Excel file
            print("\n" + "="*80)
            print("MODE: ANALYZING EXCEL FILE")
            print("="*80 + "\n")
            
            start = time.time()
            summary_text, temp_data_path, analysis_info = analyze_excel_optimized(filepath)
            
            if not summary_text:
                print("❌ Analysis failed")
                return
            
            print(f"\n✅ Analysis complete ({time.time() - start:.2f}s)")
        else:
            print(f"\n❌ Unsupported file type: {file_ext}")
            print("   Please upload .xlsx or .txt file")
            return
        
        # Load into chatbot
        print("\n" + "="*80)
        print("LOADING INTO CHATBOT")
        print("="*80 + "\n")
        
        start = time.time()
        data_info = load_data_from_temp_file(temp_data_path, summary_text)
        print(f"✅ Loaded ({time.time() - start:.2f}s)")
        
        os.unlink(temp_data_path)
        
        # Chat - simplified without example questions
        print("\n" + "="*80)
        print("CHATBOT READY - Type 'exit' to quit")
        print("="*80 + "\n")
        
        while True:
            try:
                user_input = input("🗣️  You: ").strip()
                
                if not user_input:
                    continue
                
                if user_input.lower() in ['exit', 'quit', 'q']:
                    print("\n👋 Goodbye!")
                    break
                
                print()
                start = time.time()
                answer = query_agent(user_input)
                elapsed = time.time() - start
                
                print(f"🤖 Answer ({elapsed:.2f}s):")
                print("-" * 70)
                print(answer)
                print("-" * 70 + "\n")
            
            except KeyboardInterrupt:
                print("\n\n👋 Goodbye!")
                break
            except Exception as e:
                print(f"\n❌ Error: {str(e)}\n")
    
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()


# ========================================
# RUN IT!
# ========================================

if __name__ == "__main__":
    main()